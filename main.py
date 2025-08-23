import csv
import json

from openpyxl import load_workbook

from src.search import process_bank_search


def load_json(path: str) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_csv(path: str) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def load_xlsx(path: str) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2, values_only=True)]


def filter_by_status(data: list[dict], status: str) -> list[dict]:
    return [op for op in data if op.get("status", "").upper() == status.upper()]


def sort_by_date(data: list[dict], ascending: bool = True) -> list[dict]:
    return sorted(data, key=lambda x: x.get("date", ""), reverse=not ascending)


def filter_rub(data: list[dict]) -> list[dict]:
    return [op for op in data if op.get("currency") == "RUB"]


def print_transactions(data: list[dict]):
    print(f"\nВсего банковских операций в выборке: {len(data)}\n")
    for op in data:
        print(f"{op.get('date')} {op.get('description')}")
        print(f"{op.get('from', '')} -> {op.get('to', '')}" if op.get("from") else op.get("to", ""))
        print(f"Сумма: {op.get('amount')} {op.get('currency')}\n")


def main():
    print(
        "Привет! Добро пожаловать в программу работы с банковскими транзакциями.\n"
        "Выберите необходимый пункт меню:\n"
        "1. Получить информацию о транзакциях из JSON-файла\n"
        "2. Получить информацию о транзакциях из CSV-файла\n"
        "3. Получить информацию о транзакциях из XLSX-файла"
    )

    choice = input("Ваш выбор: ")

    if choice == "1":
        print("Для обработки выбран JSON-файл.")
        data = load_json("transactions.json")
    elif choice == "2":
        print("Для обработки выбран CSV-файл.")
        data = load_csv("transactions.csv")
    elif choice == "3":
        print("Для обработки выбран XLSX-файл.")
        data = load_xlsx("transactions.xlsx")
    else:
        print("Некорректный выбор.")
        return
    valid_statuses = ["EXECUTED", "CANCELED", "PENDING"]
    while True:
        status = input(
            f"\nВведите статус, по которому необходимо выполнить фильтрацию. "
            f"\nДоступные для фильтровки статусы: {', '.join(valid_statuses)}\n"
        )
        if status.upper() in valid_statuses:
            data = filter_by_status(data, status)
            print(f'Операции отфильтрованы по статусу "{status.upper()}"')
            break
        else:
            print(f'Статус операции "{status}" недоступен.')

    if input("\nОтсортировать операции по дате? Да/Нет\n").strip().lower() == "да":
        order = input("Отсортировать по возрастанию или по убыванию?\n").strip().lower()
        ascending = order.startswith("по в")  # "по возрастанию"
        data = sort_by_date(data, ascending)

    if input("\nВыводить только рублевые транзакции? Да/Нет\n").strip().lower() == "да":
        data = filter_rub(data)

    if input("\nОтфильтровать список транзакций по определенному слову в описании? Да/Нет\n").strip().lower() == "да":
        word = input("Введите слово для поиска: ")
        data = process_bank_search(data, word)

    if not data:
        print("\nНе найдено ни одной транзакции, подходящей под ваши условия фильтрации")
    else:
        print("\nРаспечатываю итоговый список транзакций...")
        print_transactions(data)


if __name__ == "__main__":
    main()
