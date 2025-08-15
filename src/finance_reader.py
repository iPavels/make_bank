import csv
from typing import Dict, List

import openpyxl


def read_transactions_from_csv(file_path: str) -> List[Dict]:
    """
    Считывает список финансовых операций из CSV-файла.
    :param file_path: Путь к CSV-файлу.
    :return: Список словарей, где ключи — названия столбцов, значения — данные строк.
    """
    with open(file_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        return list(reader)


def read_transactions_from_excel(file_path: str) -> List[Dict]:
    """
    Считывает список финансовых операций из Excel-файла (формат .xlsx).
    :param file_path: Путь к Excel-файлу.
    :return: Список словарей, где ключи — названия столбцов (из первой строки), значения — данные строк.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    transactions = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        transaction = dict(zip(headers, row))
        transactions.append(transaction)

    return transactions
